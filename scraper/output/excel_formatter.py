"""
Excel output formatter with support for multiple sheets, formatting, and charts.
"""
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from loguru import logger

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ExcelFormatter:
    """
    Formats scraped data to Excel with multiple sheets, formatting, and styling.
    """

    def __init__(self, config: dict):
        self.config = config.get("output", {})
        self.default_sheet = "Data"
        self.freeze_panes = (1, 0)  # Freeze header row
        self.auto_filter = True

    def _flatten_record(self, record: dict, parent_key: str = "", sep: str = "_") -> dict:
        """Flatten nested dictionary."""
        items = {}

        for key, value in record.items():
            new_key = f"{parent_key}{sep}{key}" if parent_key else key

            if isinstance(value, dict):
                items.update(self._flatten_record(value, new_key, sep))
            elif isinstance(value, list):
                import json
                items[new_key] = json.dumps(value)
            elif isinstance(value, datetime):
                items[new_key] = value
            else:
                items[new_key] = value

        return items

    def _prepare_dataframe(self, data: list[dict]) -> "pd.DataFrame":
        """Convert records to pandas DataFrame."""
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas required for Excel export: pip install pandas openpyxl")

        # Flatten nested records
        flattened = [self._flatten_record(r) for r in data]

        # Create DataFrame
        df = pd.DataFrame(flattened)

        return df

    def save(
        self,
        data: list[dict],
        filepath: str,
        sheet_name: Optional[str] = None,
        create_dirs: bool = True,
    ) -> bool:
        """
        Save data to Excel file.

        Args:
            data: Data to save
            filepath: Output file path
            sheet_name: Name of the worksheet
            create_dirs: Create parent directories if needed

        Returns:
            True if successful
        """
        if not PANDAS_AVAILABLE:
            logger.error("pandas required for Excel export")
            return False

        try:
            if create_dirs:
                Path(filepath).parent.mkdir(parents=True, exist_ok=True)

            if sheet_name is None:
                sheet_name = self.default_sheet

            df = self._prepare_dataframe(data)

            # Write to Excel with formatting
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Apply formatting
                self._apply_formatting(worksheet, df)

            logger.info(f"Saved {len(data)} records to {filepath}")
            return True

        except Exception as e:
            logger.error(f"Failed to save Excel: {e}")
            return False

    def _apply_formatting(self, worksheet, df: "pd.DataFrame"):
        """Apply formatting to the worksheet."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_num, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Auto-adjust column widths
            for col_num, col_name in enumerate(df.columns, 1):
                max_length = len(str(col_name))
                for row_num in range(2, len(df) + 2):
                    cell_value = str(worksheet.cell(row=row_num, column=col_num).value or "")
                    max_length = max(max_length, len(cell_value))
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

            # Freeze header row
            if self.freeze_panes:
                worksheet.freeze_panes = worksheet.cell(
                    row=self.freeze_panes[0] + 1,
                    column=self.freeze_panes[1] + 1
                )

            # Auto filter
            if self.auto_filter:
                worksheet.auto_filter.ref = worksheet.dimensions

            # Alternating row colors
            light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for row_num in range(2, len(df) + 2):
                if row_num % 2 == 0:
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        if not cell.fill or cell.fill.fill_type == "none":
                            cell.fill = light_fill

        except ImportError:
            # Formatting not available, skip
            pass

    def save_multiple_sheets(
        self,
        sheets: dict[str, list[dict]],
        filepath: str,
        create_dirs: bool = True,
    ) -> bool:
        """
        Save multiple sheets to a single Excel file.

        Args:
            sheets: Dict mapping sheet names to data lists
            filepath: Output file path
            create_dirs: Create parent directories if needed

        Returns:
            True if successful
        """
        if not PANDAS_AVAILABLE:
            logger.error("pandas required for Excel export")
            return False

        try:
            if create_dirs:
                Path(filepath).parent.mkdir(parents=True, exist_ok=True)

            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                for sheet_name, data in sheets.items():
                    if data:
                        df = self._prepare_dataframe(data)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                        worksheet = writer.sheets[sheet_name]
                        self._apply_formatting(worksheet, df)

            logger.info(f"Saved {sum(len(d) for d in sheets.values())} records "
                       f"across {len(sheets)} sheets to {filepath}")
            return True

        except Exception as e:
            logger.error(f"Failed to save multi-sheet Excel: {e}")
            return False

    def append_sheet(
        self,
        data: list[dict],
        filepath: str,
        sheet_name: Optional[str] = None,
    ) -> bool:
        """
        Append a new sheet to existing Excel file.

        Args:
            data: Data to add
            filepath: Existing Excel file
            sheet_name: Name for the new sheet

        Returns:
            True if successful
        """
        if not PANDAS_AVAILABLE:
            logger.error("pandas required for Excel export")
            return False

        if not PANDAS_AVAILABLE:
            return False

        try:
            if sheet_name is None:
                sheet_name = self.default_sheet

            # Load existing workbook
            from openpyxl import load_workbook

            df = self._prepare_dataframe(data)

            if os.path.exists(filepath):
                with pd.ExcelWriter(filepath, engine="openpyxl", mode="a") as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            logger.info(f"Added sheet '{sheet_name}' with {len(data)} records to {filepath}")
            return True

        except Exception as e:
            logger.error(f"Failed to append Excel sheet: {e}")
            return False

    def save_with_summary(
        self,
        data: list[dict],
        filepath: str,
        summary_data: Optional[dict] = None,
        create_dirs: bool = True,
    ) -> bool:
        """
        Save data with a summary sheet.

        Args:
            data: Main data to save
            filepath: Output file path
            summary_data: Summary statistics to show
            create_dirs: Create parent directories

        Returns:
            True if successful
        """
        sheets = {"Data": data}

        if summary_data:
            # Create summary as a list of key-value pairs
            summary_rows = [{"Metric": k, "Value": v} for k, v in summary_data.items()]
            sheets["Summary"] = summary_rows

        return self.save_multiple_sheets(sheets, filepath, create_dirs)

    def load(self, filepath: str, sheet_name: Optional[str] = None) -> list[dict]:
        """
        Load data from Excel file.

        Args:
            filepath: Path to Excel file
            sheet_name: Specific sheet to load (None = first sheet)

        Returns:
            List of dictionaries
        """
        if not PANDAS_AVAILABLE:
            logger.error("pandas required for Excel loading")
            return []

        try:
            if sheet_name:
                df = pd.read_excel(filepath, sheet_name=sheet_name)
            else:
                df = pd.read_excel(filepath)

            # Convert to list of dicts
            return df.to_dict(orient="records")

        except Exception as e:
            logger.error(f"Failed to load Excel: {e}")
            return []

    def load_all_sheets(self, filepath: str) -> dict[str, list[dict]]:
        """Load all sheets from Excel file."""
        if not PANDAS_AVAILABLE:
            return {}

        try:
            sheets = pd.read_excel(filepath, sheet_name=None)
            return {
                name: df.to_dict(orient="records")
                for name, df in sheets.items()
            }
        except Exception as e:
            logger.error(f"Failed to load Excel sheets: {e}")
            return {}
